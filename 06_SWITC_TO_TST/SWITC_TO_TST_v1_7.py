import os
import re
import threading
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
import tkinter as tk
from tkinter import filedialog, messagebox
from datetime import datetime

# ==============================
# 상수 정의 및 설정 스위치
# ==============================

# [ON/OFF 스위치] S열의 값이 "수정", "추가"인 경우만 추출할지 여부 결정
# True: 수정/추가 행만 추출 | False: 전체 행 추출
ENABLE_STATUS_FILTER = True

# COMPOUND 조건 여부
ENABLE_STATUS_COMPOUND = True

# VectorCAST TST 파일 상단 고정 헤더
TST_HEADER = [
    "-- <TEST.SCRIPT>",
    "TEST.SCRIPT_FEATURE:C_DIRECT_ARRAY_INDEXING",
    "TEST.SCRIPT_FEATURE:CPP_CLASS_OBJECT_REVISION",
    "TEST.SCRIPT_FEATURE:MULTIPLE_UUT_SUPPORT",
    "TEST.SCRIPT_FEATURE:REMOVED_CL_PREFIX",
    "TEST.SCRIPT_FEATURE:MIXED_CASE_NAMES",
    "TEST.SCRIPT_FEATURE:STATIC_HEADER_FUNCS_IN_UUTS",
    "TEST.SCRIPT_FEATURE:VCAST_MAIN_NOT_RENAMED",
    "",
    "-- <TESTCASES>"
]

# stub 처리 대상 함수 목록
STUB_FUNC_LIST = [
    "OWN_LIN_Do", "LIMIT_Do", "DIAG_Do",
    "APP2MOTOR_IF_UpdateMotorInputs",
    "MOTOR_Do", "APP2MOTOR_IF_UpdateAppInputs"
]


# ==============================
# 공통 유틸 함수
# ==============================

def save_tst_file(lines, folder, filename):
    path = os.path.join(folder, filename)
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(map(str, lines)))
    return path


def save_log_file(log_entries, folder, filename, now_str):
    timestamp_display = datetime.strptime(now_str, "%Y%m%d%H%M%S").strftime("%Y-%m-%d %H:%M:%S")
    header = f"[변환 시각] {timestamp_display}"

    if not log_entries:
        contents = [header, "[결과] 변환 성공 - 실패 항목 없음"]
    else:
        contents = [
            header,
            f"[결과] 변환 실패 항목 {len(log_entries)}건",
            "",
            *log_entries,
        ]

    path = os.path.join(folder, filename)
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(contents))
    return path


def split_lines(text):
    return [line.strip() for line in str(text).splitlines() if line.strip()] if text else []


def unmerge_excel(filepath, sheetname, range_str, save_as=None, status_callback=None):
    wb = load_workbook(filepath)
    ws = wb[sheetname]

    min_c, min_r, max_c, max_r = range_boundaries(range_str)

    merged_count = 0
    for merged_cell in list(ws.merged_cells):
        if (merged_cell.min_row >= min_r and merged_cell.max_row <= max_r and
                merged_cell.min_col >= min_c and merged_cell.max_col <= max_c):
            value = ws.cell(merged_cell.min_row, merged_cell.min_col).value
            ws.unmerge_cells(
                start_row=merged_cell.min_row, start_column=merged_cell.min_col,
                end_row=merged_cell.max_row, end_column=merged_cell.max_col
            )
            for row in range(merged_cell.min_row, merged_cell.max_row + 1):
                for col in range(merged_cell.min_col, merged_cell.max_col + 1):
                    ws.cell(row=row, column=col).value = value
            merged_count += 1

    if not save_as:
        base, ext = os.path.splitext(filepath)
        save_as = f"{base}_unmerge{ext}"

    wb.save(save_as)
    if status_callback:
        status_callback(f"병합 셀 해제 완료 ({merged_count}개 범위 처리됨)\n")
    return save_as


# ==============================
# 포맷 변환 함수
# ==============================

def stub_command(func):
    if func in STUB_FUNC_LIST:
        return [f"TEST.STUB:uut_prototype_stubs.{s}"
                for s in STUB_FUNC_LIST if s != func]
    else:
        return []


def format_input_lines(input_lines):
    result_lines = []
    failures = []
    in_user_code_block = False

    for line in input_lines:
        raw_line = line.strip()
        if not raw_line:
            continue

        # 1. USER_CODE 블록 내부에 진입한 경우: 모든 라인을 그대로 추가
        if in_user_code_block:
            result_lines.append(raw_line)
            # END 태그를 만나면 블록 상태 해제
            if raw_line.startswith("TEST.END_"):
                in_user_code_block = False
            continue

        # 2. USER_CODE 시작 태그를 만난 경우 (완성형 TST)
        if raw_line.startswith("TEST.") and "_USER_CODE:" in raw_line:
            result_lines.append(raw_line)
            in_user_code_block = True
            continue

        # 3. 일반적인 완성형 TST 포맷인 경우 그대로 통과
        if raw_line.startswith("TEST."):
            result_lines.append(raw_line)
            continue

        # 4. 기존 양식 파싱 (var. / fun.)
        lower_line = raw_line.lower()

        if not lower_line.startswith(("var.", "fun.")):
            failures.append((raw_line, "Input 형식 오류: 처리 불가 형식 ('var.', 'fun.' 또는 'TEST.' 로 시작해야 함)"))
            continue

        if lower_line.startswith("fun."):
            parts = raw_line.split(":", 1)
            if len(parts) == 2:
                content, value = parts
                func_name = content.strip().split(".")[2]
                result_lines.append(f"TEST.STUB:uut_prototype_stubs.{func_name}")
                result_lines.append(f"TEST.VALUE:uut_prototype_stubs.{func_name}.return:{value.strip()}")
            else:
                failures.append((raw_line, "fun. 형식 오류: ':' 구분자 없음"))
                result_lines.append(f"TEST.VALUE:{raw_line}")
            continue

        parts = raw_line.split(":", 1)
        if len(parts) != 2:
            failures.append((raw_line, "var. 형식 오류: ':' 구분자 없음"))
            result_lines.append(f"TEST.VALUE:{raw_line}")
            continue

        content, value_expr = parts
        tokens = content.strip().split(".")
        if len(tokens) < 3:
            failures.append((raw_line, f"var. 형식 오류: 경로 토큰 부족 ({len(tokens)}개, 최소 3개 필요)"))
            result_lines.append(f"TEST.VALUE:{raw_line}")
            continue

        component = "uut_prototype_stubs" if "stub" in raw_line else tokens[1]
        full_var = f"{component}.<<GLOBAL>>." + ".".join(tokens[2:])
        value_expr = value_expr.strip()

        if "<<max>>" in value_expr or "<<min>>" in value_expr:
            result_lines.append(f"TEST.VALUE:{full_var}:{value_expr}")
            continue

        if "<<" in value_expr or ">>" in value_expr:
            prefix_var = f"{component}.<<GLOBAL>>.{tokens[2]}"
            remain = ".".join(tokens[3:])
            remain = f".{remain}" if remain else ""
            result_lines.append(f"TEST.VALUE_USER_CODE:{full_var}")
            result_lines.append(f"<<{prefix_var}>>{remain} |= ( {value_expr} );")
            result_lines.append(f"TEST.END_VALUE_USER_CODE:")
            continue

        result_lines.append(f"TEST.VALUE:{full_var}:{value_expr}")

    return result_lines, failures


def format_output_lines(output_lines):
    prefix_expected = "TEST.EXPECTED:"
    prefix_user = "TEST.EXPECTED_GLOBALS_USER_CODE:"
    suffix_user = "TEST.END_EXPECTED_GLOBALS_USER_CODE:"

    grouped_lines, normal_lines = [], []
    last_full_var, code_lines = None, []
    failures = []
    in_user_code_block = False

    for line in output_lines:
        raw_line = line.strip()
        if not raw_line:
            continue

        # 1. USER_CODE 블록 내부에 진입한 경우
        if in_user_code_block:
            normal_lines.append(raw_line)
            if raw_line.startswith("TEST.END_"):
                in_user_code_block = False
            continue

        # 2. USER_CODE 블록 시작 (완성형 TST)
        if raw_line.startswith("TEST.") and "_USER_CODE:" in raw_line:
            # 이전에 진행 중이던 그룹 로직이 있다면 닫아줌
            if last_full_var and code_lines:
                grouped_lines.append(f"{prefix_user}{last_full_var}")
                grouped_lines.extend(code_lines)
                grouped_lines.append(suffix_user)
                last_full_var = None
                code_lines = []

            normal_lines.append(raw_line)
            in_user_code_block = True
            continue

        # 3. 일반적인 완성형 TST 포맷인 경우 그대로 통과
        if raw_line.startswith("TEST."):
            if last_full_var and code_lines:
                grouped_lines.append(f"{prefix_user}{last_full_var}")
                grouped_lines.extend(code_lines)
                grouped_lines.append(suffix_user)
                last_full_var = None
                code_lines = []

            normal_lines.append(raw_line)
            continue

        # 4. 기존 양식 파싱 (var.)
        lower_line = raw_line.lower()
        if not lower_line.startswith("var."):
            failures.append((raw_line, "Output 형식 오류: 처리 불가 형식 ('var.' 또는 'TEST.' 로 시작해야 함)"))
            continue

        try:
            content, value_expr = raw_line.split(":", 1)
        except ValueError:
            failures.append((raw_line, "Output 형식 오류: ':' 구분자 없음"))
            continue

        parts = content.strip().split(".")
        component = "uut_prototype_stubs" if "stub" in raw_line else parts[1]
        value_expr = value_expr.strip()
        value_expr = value_expr.replace("<<max>>", "4294967295")

        var_parts = content.strip().split(".")
        if len(var_parts) < 3:
            failures.append((raw_line, f"Output 형식 오류: 경로 토큰 부족 ({len(var_parts)}개, 최소 3개 필요)"))
            continue

        var_path = ".".join(var_parts[2:]).strip()
        full_var = f"{component}.<<GLOBAL>>.{var_path}"

        if "&" in content:
            var_expr, bitmask = map(str.strip, content.split("&", 1))
            var_expr_parts = var_expr.split(".")
            var_path = ".".join(var_expr_parts[2:])
            full_var = f"{component}.<<GLOBAL>>.{var_path}"
            prefix = f"{component}.<<GLOBAL>>." + ".".join(var_expr_parts[2:3])
            suffix = ".".join(var_expr_parts[3:]) if len(var_expr_parts) > 3 else ""
            suffix_str = f".{suffix}" if suffix else ""
            condition = f"{{{{ (<<{prefix}>>{suffix_str} & {bitmask}) == ( {value_expr} ) }}}};"

        elif ("+" in raw_line) or ("-" in raw_line) or (not value_expr.isdigit()):
            prefix = f"{component}.<<GLOBAL>>." + ".".join(var_parts[2:3])
            suffix = ".".join(var_parts[3:]) if len(var_parts) > 3 else ""
            suffix_str = f".{suffix}" if suffix else ""
            condition = f"{{{{ (<<{prefix}>>{suffix_str}) == ( {value_expr} ) }}}};"

        else:
            normal_lines.append(f"{prefix_expected}{full_var}:{value_expr}")
            continue

        if last_full_var is None:
            last_full_var = full_var
            code_lines.append(condition)
        elif last_full_var == full_var:
            code_lines.append(condition)
        else:
            grouped_lines.append(f"{prefix_user}{last_full_var}")
            grouped_lines.extend(code_lines)
            grouped_lines.append(suffix_user)
            last_full_var = full_var
            code_lines = [condition]

    if last_full_var and code_lines:
        grouped_lines.append(f"{prefix_user}{last_full_var}")
        grouped_lines.extend(code_lines)
        grouped_lines.append(suffix_user)

    grouped_lines.extend(normal_lines)
    return grouped_lines, failures


# ==============================
# COMPOUND TC 파싱 및 생성 함수
# ==============================

def parse_compound_slots(input_text):
    slots = []
    failures = []
    for line in split_lines(input_text):
        line = line.strip()
        if not line:
            continue
        
        # 큰따옴표 안의 값들을 토큰으로 추출
        tokens = re.findall(r'"([^"]*)"', line)
        
        if len(tokens) == 5:
            seq_str, comp, func, repeat, name = tokens
            
            # 순번(seq)에 '~' 기호가 있는지 확인하여 범위 처리
            if "~" in seq_str:
                try:
                    # '~'를 기준으로 시작 번호와 끝 번호 분리
                    start_str, end_str = seq_str.split("~")
                    start_idx = int(start_str.strip())
                    end_idx = int(end_str.strip())
                    
                    # 시작 번호부터 끝 번호까지 반복
                    step = 1 if start_idx <= end_idx else -1
                    for i in range(start_idx, end_idx + step, step):
                        slots.append([str(i), comp, func, repeat, name])
                        
                except ValueError:
                    reason = f"COMPOUND 슬롯 범위 형식 오류: 숫자 변환 실패 ('{seq_str}')"
                    failures.append((line, reason))
            else:
                # '~'가 없는 일반 단일 순번인 경우 그대로 추가
                slots.append(tokens)
        else:
            reason = f"COMPOUND 슬롯 형식 오류: 토큰 수 불일치 ({len(tokens)}개, 5개 필요)"
            failures.append((line, reason))
            
    return slots, failures


def build_compound_block(tc_id, slots):
    block = [
        "TEST.SUBPROGRAM:<<COMPOUND>>",
        "TEST.NEW",
        f"TEST.NAME:{tc_id}",
    ]
    for slot in slots:
        seq, comp, func, repeat, name = slot
        block.append(f'TEST.SLOT: "{seq}", "{comp}", "{func}", "{repeat}", "{name}"')
    block.append("TEST.END")
    block.append("")
    return block


# ==============================
# 메인 로직
# ==============================

def excel_to_tst(filepath, sheetname, range_str, save_dir,
                 enable_filter=None, enable_compound=None, status_callback=None):
    use_filter = ENABLE_STATUS_FILTER if enable_filter is None else enable_filter
    use_compound = ENABLE_STATUS_COMPOUND if enable_compound is None else enable_compound

    wb = load_workbook(filepath)
    ws = wb[sheetname]
    min_c, min_r, max_c, max_r = range_boundaries(range_str)

    # COMPOUND 처리
    compound_only_tcs = set()
    for row in ws.iter_rows(min_row=min_r, max_row=max_r):
        values = [cell.value for cell in row]
        # (values[2] = TC ID), (values[9] = Inputs)
        if len(values) > 9 and values[2] is not None:
            tc_id_temp = str(values[2]).strip()
            if tc_id_temp.endswith("_C"):
                # COMPOUND TC인 경우 Input을 파싱하여 SLOT 이름 수집
                slots, _ = parse_compound_slots(values[9])
                for slot in slots:
                    if len(slot) == 5:
                        slot_tc_name = slot[4].strip()
                        compound_only_tcs.add(slot_tc_name)

    normal_lines = TST_HEADER[:]
    compound_lines = []
    log_entries = []
    tc_count = 0

    for row in ws.iter_rows(min_row=min_r, max_row=max_r):
        values = [cell.value for cell in row]

        try:
            req_id, tc_id, comp, func, pre, inputs, outputs = (
                values[1], values[2], values[6], values[7], values[8], values[9], values[10]
            )
        except IndexError:
            continue

        if not all([req_id, tc_id]):
            continue

        current_row_idx = row[0].row
        status = ws.cell(row=current_row_idx, column=19).value

        if use_filter:
            if str(status).strip() not in ["수정", "추가", "400W수정"]:
                continue

        tc_id_str = str(tc_id).strip()

        # ── COMPOUND TC 처리 ──
        if tc_id_str.endswith("_C"):
            slots, slot_failures = parse_compound_slots(inputs)
            for raw_val, reason in slot_failures:
                log_entries.append(
                    f"[행 {current_row_idx}] TC: {tc_id_str} | 항목: COMPOUND 슬롯 | 값: \"{raw_val}\" | 이유: {reason}"
                )
            if not slots:
                log_entries.append(
                    f"[행 {current_row_idx}] TC: {tc_id_str} | 항목: COMPOUND 슬롯 | 값: (없음) | 이유: 유효한 슬롯이 없어 TC 블록 생성 건너뜀"
                )
                continue
            compound_lines.extend(build_compound_block(tc_id_str, slots))
            tc_count += 1
            if status_callback:
                status_callback(f"  [COMPOUND] {tc_id_str}\n")
            continue

        # ── 일반 TC 처리 ──
        if not all([comp, func]):
            continue

        if inputs is None:
            inputs = ""
        if pre is not None:
            inputs = pre + "\n" + inputs

        # ---- 수정된 Input 파싱 로직 적용 ----
        converted_input_lines, input_failures = format_input_lines(split_lines(inputs))
        for raw_val, reason in input_failures:
            log_entries.append(
                f"[행 {current_row_idx}] TC: {tc_id_str} | 항목: Input | 값: \"{raw_val}\" | 이유: {reason}"
            )

        # ---- 수정된 Output 파싱 로직 적용 ----
        converted_output_lines, output_failures = format_output_lines(split_lines(outputs))
        for raw_val, reason in output_failures:
            log_entries.append(
                f"[행 {current_row_idx}] TC: {tc_id_str} | 항목: Output | 값: \"{raw_val}\" | 이유: {reason}"
            )

        # ---- 일반 TC 블록 생성 ----
        tc_block = [
            f"TEST.UNIT:{comp}",
            f"TEST.SUBPROGRAM:{func}",
            "TEST.NEW",
            f"TEST.NAME:{tc_id_str}"
        ]

        # COMPOUND_ONLY 추가
        if use_compound and tc_id_str in compound_only_tcs:
            tc_block.append("TEST.COMPOUND_ONLY")

        tc_block.extend(converted_input_lines)
        tc_block.extend(converted_output_lines)
        tc_block.extend(["TEST.END", ""])

        normal_lines.extend(tc_block)
        
        tc_count += 1
        if status_callback:
            status_callback(f"  [TC] {tc_id_str}\n")

    all_lines = normal_lines + compound_lines

    if len(all_lines) > len(TST_HEADER):
        now_str = datetime.now().strftime("%Y%m%d%H%M%S")
        base_name = f"SWITC_{now_str}"
        tst_path = save_tst_file(all_lines, save_dir, f"{base_name}.tst")
        log_path = save_log_file(log_entries, save_dir, f"{base_name}.log", now_str)
        return tst_path, log_path, tc_count, len(log_entries)
    else:
        return None, None, 0, 0


# ==============================
# GUI
# ==============================

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("SWITC → TST 변환기")
        self.geometry("740x620")
        self.resizable(False, False)

        # ── 로즈 에스프레소 다크 팔레트 ──
        self.BG      = "#1f1a1a"
        self.BG2     = "#2a2020"
        self.BORDER  = "#5a3a3a"
        self.ACCENT  = "#f08080"
        self.ACCENT2 = "#7dcea0"
        self.FG      = "#f5e6e6"
        self.FG_DIM  = "#b08880"
        self.BTN_BG  = "#3d2828"
        self.BTN_HOV = "#543838"
        self.RUN_BG  = "#c0504a"
        self.RUN_HOV = "#a84040"
        self.WARN    = "#f9e2af"

        self.configure(bg=self.BG)

        self._excel_path  = tk.StringVar(value="선택된 파일 없음")
        self._output_dir  = tk.StringVar(value="")
        self._filter_var  = tk.BooleanVar(value=ENABLE_STATUS_FILTER)
        self._compound_var  = tk.BooleanVar(value=ENABLE_STATUS_COMPOUND)

        self._build_ui()

    def _build_ui(self):
        header = tk.Frame(self, bg=self.BG)
        header.pack(fill="x", padx=32, pady=(28, 0))

        tk.Label(
            header,
            text="SWITC  →  TST",
            font=("Segoe UI", 20, "bold"),
            fg=self.ACCENT,
            bg=self.BG,
        ).pack(side="left")

        tk.Label(
            header,
            text="엑셀 → VectorCAST TST 변환 도구",
            font=("Segoe UI", 9),
            fg=self.FG_DIM,
            bg=self.BG,
        ).pack(side="left", padx=(14, 0), pady=(8, 0))

        tk.Frame(self, bg=self.BORDER, height=1).pack(fill="x", padx=32, pady=16)

        self._section("엑셀 파일 선택")
        file_row = tk.Frame(self, bg=self.BG)
        file_row.pack(fill="x", padx=32, pady=(4, 0))

        tk.Label(
            file_row,
            textvariable=self._excel_path,
            font=("Segoe UI", 9),
            fg=self.FG,
            bg=self.BG2,
            anchor="w",
            relief="flat",
            padx=12,
            pady=7,
            width=52,
        ).pack(side="left")

        self._btn(file_row, "파일 선택", self._choose_file).pack(side="left", padx=(8, 0))

        self._section("출력 폴더")
        dir_row = tk.Frame(self, bg=self.BG)
        dir_row.pack(fill="x", padx=32, pady=(4, 0))

        tk.Label(
            dir_row,
            textvariable=self._output_dir,
            font=("Segoe UI", 9),
            fg=self.FG,
            bg=self.BG2,
            anchor="w",
            relief="flat",
            padx=12,
            pady=7,
            width=52,
        ).pack(side="left")

        self._btn(dir_row, "폴더 선택", self._choose_dir).pack(side="left", padx=(8, 0))

        self._section("변환 옵션")
        opt_row = tk.Frame(self, bg=self.BG)
        opt_row.pack(fill="x", padx=32, pady=(4, 0))

        filter_cb = tk.Checkbutton(
            opt_row,
            text='S열 필터 활성화  (상태값이 "수정" 또는 "추가"인 행만 변환)',
            variable=self._filter_var,
            font=("Segoe UI", 9),
            fg=self.FG,
            bg=self.BG,
            activebackground=self.BG,
            activeforeground=self.FG,
            selectcolor=self.BTN_BG,
            relief="flat",
            cursor="hand2",
        )
        filter_cb.pack(anchor="w", pady=(0, 6))

        # COMPOUND 체크박스
        compound_cb = tk.Checkbutton(
            opt_row,
            text='TEST.COMPOUND 추가 (COMPOUND 테스트의 요소일 경우 조건 옵션 추가)',
            variable=self._compound_var,
            font=("Segoe UI", 9),
            fg=self.FG,
            bg=self.BG,
            activebackground=self.BG,
            activeforeground=self.FG,
            selectcolor=self.BTN_BG,
            relief="flat",
            cursor="hand2",
        )
        compound_cb.pack(anchor="w")

        tk.Frame(self, bg=self.BORDER, height=1).pack(fill="x", padx=32, pady=16)

        self._run_btn = tk.Button(
            self,
            text="▶   변환 실행",
            font=("Segoe UI", 11, "bold"),
            fg="#ffffff",
            bg=self.RUN_BG,
            activebackground=self.RUN_HOV,
            activeforeground="#ffffff",
            relief="flat",
            cursor="hand2",
            padx=28,
            pady=10,
            command=self._run,
        )
        self._run_btn.pack(padx=32, anchor="w")

        self._section("실행 로그")
        log_frame = tk.Frame(self, bg=self.BG)
        log_frame.pack(fill="both", expand=True, padx=32, pady=(4, 24))

        self._log = tk.Text(
            log_frame,
            font=("Consolas", 9),
            fg=self.FG,
            bg=self.BG2,
            insertbackground=self.FG,
            selectbackground=self.BTN_HOV,
            relief="flat",
            state="disabled",
            wrap="word",
            bd=0,
            padx=12,
            pady=10,
        )
        scroll = tk.Scrollbar(log_frame, command=self._log.yview, bg=self.BG2,
                              troughcolor=self.BG2, relief="flat")
        self._log.configure(yscrollcommand=scroll.set)
        scroll.pack(side="right", fill="y")
        self._log.pack(fill="both", expand=True)

        self._log.tag_configure("ok",   foreground=self.ACCENT2)
        self._log.tag_configure("err",  foreground="#f28b82")
        self._log.tag_configure("warn", foreground=self.WARN)
        self._log.tag_configure("hd",   foreground=self.ACCENT)

    def _section(self, text):
        tk.Label(
            self,
            text=text,
            font=("Segoe UI", 8, "bold"),
            fg=self.ACCENT,
            bg=self.BG,
        ).pack(anchor="w", padx=32, pady=(12, 2))

    def _btn(self, parent, text, cmd):
        return tk.Button(
            parent,
            text=text,
            font=("Segoe UI", 9),
            fg=self.FG,
            bg=self.BTN_BG,
            activebackground=self.BTN_HOV,
            activeforeground=self.FG,
            relief="flat",
            cursor="hand2",
            padx=14,
            pady=6,
            command=cmd,
        )

    def _choose_file(self):
        path = filedialog.askopenfilename(
            title="SWITC 엑셀 파일 선택",
            filetypes=[("Excel 파일", "*.xlsx *.xls"), ("모든 파일", "*.*")],
        )
        if path:
            self._excel_path.set(path)
            if not self._output_dir.get():
                self._output_dir.set(os.path.dirname(path))
            self._log_write(f"파일 선택됨: {os.path.basename(path)}\n", tag="hd")

            try:
                wb_temp = load_workbook(path, read_only=True)
                sheets = wb_temp.sheetnames
                wb_temp.close()
                self._log_write(f"시트 목록: {', '.join(sheets)}\n")

                target = None
                if "Requirements TestCase" in sheets:
                    target = "Requirements TestCase"
                elif "TestCase" in sheets:
                    target = "TestCase"

                if target:
                    self._log_write(f"대상 시트: {target}\n", tag="ok")
                else:
                    self._log_write(
                        "경고: 'Requirements TestCase' 또는 'TestCase' 시트를 찾을 수 없습니다.\n",
                        tag="warn"
                    )
            except Exception as e:
                self._log_write(f"[오류] 파일 미리보기 실패: {e}\n", tag="err")

    def _choose_dir(self):
        path = filedialog.askdirectory(title="출력 폴더 선택")
        if path:
            self._output_dir.set(path)
            self._log_write(f"출력 폴더: {path}\n")

    def _run(self):
        src = self._excel_path.get()
        out = self._output_dir.get()

        if src == "선택된 파일 없음" or not os.path.isfile(src):
            messagebox.showerror("오류", "엑셀 파일을 먼저 선택하세요.")
            return
        if not out or not os.path.isdir(out):
            messagebox.showerror("오류", "유효한 출력 폴더를 선택하세요.")
            return

        self._run_btn.configure(state="disabled", text="⏳  변환 중...")
        thread = threading.Thread(target=self._run_worker, args=(src, out), daemon=True)
        thread.start()

    def _run_worker(self, src, out):
        self._log_write("\n── 변환 시작 ──\n", tag="hd")
        rng = "B10:S10000"

        try:
            wb_temp = load_workbook(src, read_only=True)
            available_sheets = wb_temp.sheetnames
            wb_temp.close()

            target_sheet = None
            if "Requirements TestCase" in available_sheets:
                target_sheet = "Requirements TestCase"
            elif "TestCase" in available_sheets:
                target_sheet = "TestCase"

            if not target_sheet:
                self._log_write(
                    "[오류] 'Requirements TestCase' 또는 'TestCase' 시트를 찾을 수 없습니다.\n",
                    tag="err"
                )
                self._finish_btn()
                return

            self._log_write(f"대상 시트: {target_sheet}\n")

            use_filter = self._filter_var.get()
            if use_filter:
                self._log_write("S열 필터 활성화: '수정', '추가' 행만 추출\n", tag="warn")

            # COMPOUND 조건
            use_compound = self._compound_var.get()
            if use_compound:
                self._log_write("COMPOUND 옵션 추가\n", tag="warn")

            self._log_write("병합 셀 해제 중...\n")
            unmerge_path = unmerge_excel(
                src, target_sheet, rng,
                status_callback=self._log_write
            )

            self._log_write("TST 변환 중...\n")
            tst_path, log_path, tc_count, fail_count = excel_to_tst(
                unmerge_path, target_sheet, rng, out,
                enable_filter=use_filter,
                enable_compound=use_compound,
                status_callback=self._log_write
            )

            if tst_path:
                self._log_write(f"\n✔  완료: {tc_count}개 TC 변환됨\n", tag="ok")
                self._log_write(f"   TST 파일: {os.path.basename(tst_path)}\n")
                self._log_write(f"   LOG 파일: {os.path.basename(log_path)}\n")
                self._log_write(f"   저장 위치: {out}\n")

                if fail_count:
                    self._log_write(f"   ⚠  변환 실패 항목: {fail_count}건 (로그 파일 참조)\n", tag="warn")

                summary = (
                    f"변환 완료!\n\n"
                    f"생성된 TC: {tc_count}개\n"
                    f"실패 항목: {fail_count}건\n\n"
                    f"저장 위치:\n{out}"
                )
                self.after(0, lambda: messagebox.showinfo("완료", summary))
            else:
                self._log_write(
                    "\n[경고] 변환된 TC가 없습니다. 필터 설정 또는 원본 데이터를 확인하세요.\n",
                    tag="warn"
                )
                self.after(0, lambda: messagebox.showwarning(
                    "결과 없음",
                    "변환된 TC가 없습니다.\nS열 필터 설정 또는 원본 데이터를 확인하세요."
                ))

        except Exception as e:
            self._log_write(f"[오류] {e}\n", tag="err")
            self.after(0, lambda: messagebox.showerror("오류", str(e)))

        finally:
            self._finish_btn()

    def _finish_btn(self):
        self.after(0, lambda: self._run_btn.configure(state="normal", text="▶   변환 실행"))

    def _log_write(self, text, tag=None):
        def _write():
            self._log.configure(state="normal")
            if tag:
                self._log.insert("end", text, tag)
            else:
                self._log.insert("end", text)
            self._log.see("end")
            self._log.configure(state="disabled")
        self.after(0, _write)


if __name__ == "__main__":
    app = App()
    app.mainloop()